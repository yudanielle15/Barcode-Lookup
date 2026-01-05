import streamlit as st
import pandas as pd
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

# -------------------------------
# Page config
# -------------------------------
st.set_page_config(page_title="Biomarker Barcode Scanner", layout="centered")
st.title("🔬 Biomarker Sample Barcode Scanner")
st.write("Scan or type barcodes → they become removable bubbles → process all at once")
st.divider()

# -------------------------------
# Initialize session state
# -------------------------------
defaults = {
    "df": None,
    "barcode_tags": [],
    "matched_df": pd.DataFrame(),
    "unmatched_barcodes": [],
    "barcode_input": "",
    "uploaded_file_name": None  # Track uploaded file
}

for key, value in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value

# -------------------------------
# Upload Excel
# -------------------------------
uploaded_file = st.file_uploader("📁 Upload your sample Excel file", type=["xlsx"])

def load_file(file):
    try:
        df = pd.read_excel(file)
    except Exception as e:
        st.error(f"❌ Failed to read Excel: {e}")
        return None

    if "Barcode" not in df.columns:
        st.error("❌ Excel must contain a 'Barcode' column.")
        return None

    df["Scan_Status"] = df.get("Scan_Status", "")
    df["Barcode"] = df["Barcode"].astype(str)

    # Update DataFrame and reset processed info, but keep scanned barcodes
    st.session_state.df = df
    st.session_state.matched_df = pd.DataFrame()
    st.session_state.unmatched_barcodes = []
    st.session_state.barcode_input = ""

    st.success("✅ File loaded. Ready to scan.")
    return df

if uploaded_file:
    # Only reset session state if the file is different
    if st.session_state.uploaded_file_name != uploaded_file.name:
        load_file(uploaded_file)
        st.session_state.uploaded_file_name = uploaded_file.name

st.divider()

# -------------------------------
# Barcode input
# -------------------------------
st.subheader("🧪 Scan / Type Barcodes")

def add_barcode():
    barcode = st.session_state.barcode_input.strip()
    if barcode and barcode not in st.session_state.barcode_tags:
        st.session_state.barcode_tags.append(barcode)
    st.session_state.barcode_input = ""

st.text_input("Type or scan barcode", key="barcode_input", on_change=add_barcode)

# -------------------------------
# Display scanned barcodes (removable)
# -------------------------------
if st.session_state.barcode_tags:
    selected = st.multiselect(
        "Scanned barcodes (click ❌ to remove):",
        options=st.session_state.barcode_tags,
        default=st.session_state.barcode_tags
    )
    st.session_state.barcode_tags = selected

st.divider()

# -------------------------------
# Process all barcodes
# -------------------------------
if st.button("🚀 Process All Barcodes", use_container_width=True):
    if st.session_state.df is None:
        st.warning("⚠️ Upload an Excel file first.")
    else:
        barcode_list = st.session_state.barcode_tags
        df_barcodes = st.session_state.df["Barcode"].tolist()

        matched = [b for b in barcode_list if b in df_barcodes]
        unmatched = [b for b in barcode_list if b not in df_barcodes]

        st.session_state.df.loc[st.session_state.df["Barcode"].isin(matched), "Scan_Status"] = "Matched"
        st.session_state.matched_df = st.session_state.df[st.session_state.df["Barcode"].isin(matched)]
        st.session_state.unmatched_barcodes = unmatched
        st.session_state.barcode_tags = []

        st.success(f"✅ {len(matched)} matched | ❌ {len(unmatched)} unmatched")

st.divider()

# -------------------------------
# Show results
# -------------------------------
if not st.session_state.matched_df.empty:
    st.subheader("🔹 Matched Samples")
    st.dataframe(
        st.session_state.matched_df.style.apply(
            lambda row: ["background-color: yellow" if col in ["Screen ID", "Visit", "Sample Name"] else "" for col in row.index],
            axis=1
        ),
        use_container_width=True
    )

if st.session_state.unmatched_barcodes:
    st.subheader("❌ Unmatched Barcodes")
    st.code("\n".join(st.session_state.unmatched_barcodes))

st.divider()

# -------------------------------
# Download updated Excel (preserve original formatting)
# -------------------------------
if uploaded_file and st.session_state.df is not None:
    # Load original workbook to preserve formatting
    wb = load_workbook(uploaded_file)
    ws = wb.active  # assume first sheet; adjust if needed

    # Mapping Barcode -> Scan_Status
    status_map = st.session_state.df.set_index("Barcode")["Scan_Status"].to_dict()

    # Find column indices
    barcode_col = None
    scan_status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Barcode":
            barcode_col = idx
        elif cell.value == "Scan_Status":
            scan_status_col = idx

    # If Scan_Status column doesn't exist, add it at the end
    if scan_status_col is None:
        scan_status_col = ws.max_column + 1
        ws.cell(row=1, column=scan_status_col, value="Scan_Status")

    # Update Scan_Status based on barcode matches (preserve all formatting)
    for row in ws.iter_rows(min_row=2):
        barcode = str(row[barcode_col - 1].value)
        if barcode in status_map:
            row[scan_status_col - 1].value = status_map[barcode]

    # Save workbook to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    new_filename = Path(uploaded_file.name).stem + "_Scanned.xlsx"
    st.download_button(
        "💾 Download Updated Excel (Formatting Preserved)",
        buffer,
        file_name=new_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
else:
    st.info("⬆️ Upload an Excel file to begin.")
